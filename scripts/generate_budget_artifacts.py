#!/usr/bin/env python3
"""
Generate Budget Planning Artifacts (Folder 08)

Creates Excel-based budget planning tools based on structured data (folders 01-07):
1. Store-type-specific budget plans
2. Reusable configuration templates
3. Build strategy worksheets
4. Agent integration tools
"""

import os
import json
import csv
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from datetime import datetime

# Create output directory
output_dir = 'output/08_Budget_Artifacts'
os.makedirs(f'{output_dir}/budget_plans', exist_ok=True)
os.makedirs(f'{output_dir}/templates', exist_ok=True)
os.makedirs(f'{output_dir}/strategy_worksheets', exist_ok=True)
os.makedirs(f'{output_dir}/agent_tools', exist_ok=True)

# Load source data
print("Loading source data...")
with open('output/01_Build_Templates/store_types.json') as f:
    store_types_data = json.load(f)

with open('output/01_Build_Templates/base_template.json') as f:
    base_template = json.load(f)

with open('output/03_Historical_Projects/historical_projects.json') as f:
    historical_projects = json.load(f)

with open('output/04_Regional_Modifiers/regional_modifiers.json') as f:
    regional_modifiers = json.load(f)

with open('output/05_Cost_Models/cost_model_suburban_standard.json') as f:
    cost_model = json.load(f)

with open('output/06_Vendor_Data/vendor_catalog.json') as f:
    vendor_data = json.load(f)

# Styling helpers
def create_header_style():
    """Create header cell style."""
    return {
        'font': Font(bold=True, color='FFFFFF', size=12),
        'fill': PatternFill(start_color='366092', end_color='366092', fill_type='solid'),
        'alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'border': Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    }

def create_input_style():
    """Create input cell style (blue)."""
    return {
        'fill': PatternFill(start_color='D6DCE4', end_color='D6DCE4', fill_type='solid'),
        'font': Font(bold=True),
        'alignment': Alignment(horizontal='left', vertical='center')
    }

def create_output_style():
    """Create output cell style (green)."""
    return {
        'fill': PatternFill(start_color='C6E0B4', end_color='C6E0B4', fill_type='solid'),
        'font': Font(bold=True),
        'alignment': Alignment(horizontal='right', vertical='center')
    }

def create_formula_style():
    """Create formula cell style (gray)."""
    return {
        'fill': PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid'),
        'alignment': Alignment(horizontal='right', vertical='center')
    }

def apply_style(cell, style_dict):
    """Apply style dictionary to cell."""
    for key, value in style_dict.items():
        setattr(cell, key, value)

# ============================================================================
# 1. STORE-TYPE BUDGET PLANS
# ============================================================================
print("\n[1/4] Generating store-type budget plans...")

def create_budget_plan(store_type):
    """Create complete budget plan for a store type."""
    store_info = next(s for s in store_types_data['store_types'] if s['type_id'] == store_type)

    # Get historical average for this store type
    historical_avg = calculate_historical_average(store_type)

    wb = Workbook()
    wb.remove(wb.active)  # Remove default sheet

    # Worksheet 1: Executive Summary
    ws_summary = wb.create_sheet("Executive Summary")
    create_executive_summary(ws_summary, store_info, historical_avg)

    # Worksheet 2: Detailed Line Items
    ws_details = wb.create_sheet("Detailed Line Items")
    create_detailed_line_items(ws_details, store_info, historical_avg)

    # Worksheet 3: Scenario Comparisons
    ws_scenarios = wb.create_sheet("Scenario Comparisons")
    create_scenario_comparisons(ws_scenarios, store_info, historical_avg)

    # Worksheet 4: Data Sources
    ws_sources = wb.create_sheet("Data Sources")
    create_data_sources(ws_sources, store_info)

    # Worksheet 5: Agent Instructions
    ws_agent = wb.create_sheet("Agent Instructions")
    create_agent_instructions(ws_agent, store_info)

    # Save
    filename = f"{output_dir}/budget_plans/Budget_Plan_{store_type}_{store_info['typical_sqft']}sqft.xlsx"
    wb.save(filename)
    print(f"  ✓ Created {filename}")

    return filename

def calculate_historical_average(store_type):
    """Calculate average costs from historical projects."""
    projects = [p for p in historical_projects['projects'] if p['store_type'] == store_type]

    if not projects:
        return None

    avg_total = sum(p['total_cost'] for p in projects) / len(projects)
    avg_psf = sum(p['cost_per_sqft'] for p in projects) / len(projects)

    categories = {}
    for cat in ['construction', 'electrical', 'hvac', 'plumbing', 'fixtures', 'technology', 'soft_costs']:
        avg = sum(p['categories'][cat] for p in projects) / len(projects)
        categories[cat] = int(avg)

    return {
        'count': len(projects),
        'total_cost': int(avg_total),
        'cost_per_sqft': round(avg_psf, 2),
        'categories': categories
    }

def create_executive_summary(ws, store_info, historical_avg):
    """Create executive summary worksheet."""
    # Header
    ws['A1'] = f"Budget Plan: {store_info['name']}"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')

    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d')}"
    ws['A2'].font = Font(italic=True)
    ws.merge_cells('A2:D2')

    # Store configuration
    row = 4
    ws[f'A{row}'] = "STORE CONFIGURATION"
    apply_style(ws[f'A{row}'], create_header_style())
    ws.merge_cells(f'A{row}:D{row}')

    row += 1
    config_data = [
        ["Store Type:", store_info['name']],
        ["Typical Square Footage:", f"{store_info['typical_sqft']:,} sqft"],
        ["Size Range:", f"{store_info['sqft_range'][0]:,} - {store_info['sqft_range'][1]:,} sqft"],
        ["Primary Markets:", ", ".join(store_info['target_markets'][:3])],
        ["Use Cases:", ", ".join(store_info['typical_use_cases'])]
    ]

    for label, value in config_data:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        ws.merge_cells(f'B{row}:D{row}')
        row += 1

    # Cost summary
    row += 2
    ws[f'A{row}'] = "COST SUMMARY"
    apply_style(ws[f'A{row}'], create_header_style())
    ws.merge_cells(f'A{row}:D{row}')

    if historical_avg:
        row += 1
        ws[f'A{row}'] = "Category"
        ws[f'B{row}'] = "Amount"
        ws[f'C{row}'] = "% of Total"
        ws[f'D{row}'] = "Cost/SqFt"
        for col in ['A', 'B', 'C', 'D']:
            apply_style(ws[f'{col}{row}'], create_header_style())

        total = historical_avg['total_cost']
        sqft = store_info['typical_sqft']

        row += 1
        categories_display = [
            ('Construction', historical_avg['categories']['construction']),
            ('Electrical', historical_avg['categories']['electrical']),
            ('HVAC', historical_avg['categories']['hvac']),
            ('Plumbing', historical_avg['categories']['plumbing']),
            ('Fixtures', historical_avg['categories']['fixtures']),
            ('Technology', historical_avg['categories']['technology']),
            ('Soft Costs', historical_avg['categories']['soft_costs'])
        ]

        for cat_name, cat_cost in categories_display:
            ws[f'A{row}'] = cat_name
            ws[f'B{row}'] = cat_cost
            ws[f'B{row}'].number_format = '$#,##0'
            ws[f'C{row}'] = cat_cost / total
            ws[f'C{row}'].number_format = '0.0%'
            ws[f'D{row}'] = cat_cost / sqft
            ws[f'D{row}'].number_format = '$#,##0.00'
            row += 1

        # Total
        ws[f'A{row}'] = "TOTAL PROJECT COST"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = total
        ws[f'B{row}'].number_format = '$#,##0'
        ws[f'B{row}'].font = Font(bold=True)
        apply_style(ws[f'B{row}'], create_output_style())
        ws[f'D{row}'] = historical_avg['cost_per_sqft']
        ws[f'D{row}'].number_format = '$#,##0.00'
        ws[f'D{row}'].font = Font(bold=True)

        row += 2
        ws[f'A{row}'] = f"Based on {historical_avg['count']} historical projects"
        ws[f'A{row}'].font = Font(italic=True)
        ws.merge_cells(f'A{row}:D{row}')

    # Set column widths
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 15

def create_detailed_line_items(ws, store_info, historical_avg):
    """Create detailed line items worksheet."""
    # Use suburban_standard cost model as template
    model = cost_model['base_costs']['suburban_standard_3500sqft']

    # Header
    ws['A1'] = f"Detailed Line Items: {store_info['name']}"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:G1')

    row = 3
    headers = ['Category', 'Item', 'Quantity', 'Unit', 'Unit Cost', 'Total', 'Source']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, create_header_style())

    row += 1

    # Iterate through categories
    for category, cat_data in model['breakdown'].items():
        # Category header
        ws.cell(row=row, column=1, value=category.upper())
        ws.cell(row=row, column=1).font = Font(bold=True, size=11)
        ws.merge_cells(f'A{row}:G{row}')
        ws.cell(row=row, column=1).fill = PatternFill(start_color='E7E6E6', end_color='E7E6E6', fill_type='solid')
        row += 1

        # Line items
        for item in cat_data['line_items']:
            ws.cell(row=row, column=1, value='')
            ws.cell(row=row, column=2, value=item['item'])
            ws.cell(row=row, column=3, value=item['quantity'])
            ws.cell(row=row, column=4, value=item['unit'])
            ws.cell(row=row, column=5, value=item['unit_cost'])
            ws.cell(row=row, column=5).number_format = '$#,##0.00'
            ws.cell(row=row, column=6, value=item['total'])
            ws.cell(row=row, column=6).number_format = '$#,##0'
            ws.cell(row=row, column=7, value="Cost Model: suburban_standard")
            ws.cell(row=row, column=7).font = Font(size=9, italic=True)
            row += 1

        # Category subtotal
        ws.cell(row=row, column=2, value=f"{category.title()} Subtotal")
        ws.cell(row=row, column=2).font = Font(bold=True)
        ws.cell(row=row, column=6, value=cat_data['total'])
        ws.cell(row=row, column=6).number_format = '$#,##0'
        ws.cell(row=row, column=6).font = Font(bold=True)
        apply_style(ws.cell(row=row, column=6), create_output_style())
        row += 2

    # Set column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 30

def create_scenario_comparisons(ws, store_info, historical_avg):
    """Create scenario comparison worksheet."""
    ws['A1'] = "Budget Scenarios Comparison"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:E1')

    if not historical_avg:
        ws['A3'] = "No historical data available for scenarios"
        return

    base_cost = historical_avg['total_cost']

    # Scenarios
    scenarios = [
        {
            'name': 'Base Case',
            'description': 'Standard build, Columbus market, 12-week timeline',
            'multiplier': 1.0,
            'notes': 'Historical average'
        },
        {
            'name': 'Cincinnati Market',
            'description': 'Union labor requirements',
            'multiplier': 1.06,
            'notes': 'Regional modifier: electrical 1.08x, construction 1.05x'
        },
        {
            'name': 'Accelerated Schedule',
            'description': '8-week timeline (vs. 12 weeks)',
            'multiplier': 1.15,
            'notes': 'Labor premium, expedite fees'
        },
        {
            'name': 'Premium Finishes',
            'description': 'Upgraded materials and fixtures',
            'multiplier': 1.22,
            'notes': 'Fixture upgrade 30%, material upgrade 15%'
        },
        {
            'name': 'Budget Cap (Value Eng.)',
            'description': 'Cost-optimized specifications',
            'multiplier': 0.88,
            'notes': 'Material substitutions, simplified finishes'
        }
    ]

    row = 3
    headers = ['Scenario', 'Description', 'Total Cost', 'vs. Base', 'Notes']
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col_idx, value=header)
        apply_style(cell, create_header_style())

    row += 1

    for scenario in scenarios:
        cost = int(base_cost * scenario['multiplier'])
        variance = scenario['multiplier'] - 1.0

        ws.cell(row=row, column=1, value=scenario['name'])
        ws.cell(row=row, column=1).font = Font(bold=True)
        ws.cell(row=row, column=2, value=scenario['description'])
        ws.cell(row=row, column=3, value=cost)
        ws.cell(row=row, column=3).number_format = '$#,##0'
        ws.cell(row=row, column=4, value=variance)
        ws.cell(row=row, column=4).number_format = '+0.0%;-0.0%;0%'

        # Color code variance
        if variance > 0:
            ws.cell(row=row, column=4).font = Font(color='C00000')  # Red for over
        elif variance < 0:
            ws.cell(row=row, column=4).font = Font(color='00B050')  # Green for under

        ws.cell(row=row, column=5, value=scenario['notes'])
        ws.cell(row=row, column=5).font = Font(size=9, italic=True)
        row += 1

    # Set column widths
    ws.column_dimensions['A'].width = 22
    ws.column_dimensions['B'].width = 35
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 40

def create_data_sources(ws, store_info):
    """Create data sources worksheet."""
    ws['A1'] = "Data Sources & References"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:C1')

    row = 3
    ws[f'A{row}'] = "Source Type"
    ws[f'B{row}'] = "Location"
    ws[f'C{row}'] = "Description"
    for col in ['A', 'B', 'C']:
        apply_style(ws[f'{col}{row}'], create_header_style())

    sources = [
        ("Store Type Definition", "01_Build_Templates/store_types.json", f"type_id: {store_info['type_id']}"),
        ("Base Template", "01_Build_Templates/base_template.json", "Category structure and specifications"),
        ("Historical Projects", "03_Historical_Projects/historical_projects.csv", f"Filtered by store_type = {store_info['type_id']}"),
        ("Cost Model", "05_Cost_Models/cost_model_suburban_standard.json", "Detailed line item costs and formulas"),
        ("Regional Modifiers", "04_Regional_Modifiers/regional_modifiers.csv", "Market-specific multipliers"),
        ("Vendor Pricing", "06_Vendor_Data/vendor_pricing.csv", "Current vendor unit costs"),
        ("Meeting Transcripts", "07_Conversations/meeting_transcripts/vendor_negotiation/*", "Pricing discussions and negotiations"),
        ("Teams Conversations", "07_Conversations/teams_channels/construction-vendors.json", "Vendor performance and cost insights")
    ]

    row += 1
    for source_type, location, description in sources:
        ws[f'A{row}'] = source_type
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = location
        ws[f'B{row}'].font = Font(name='Courier New', size=9)
        ws[f'C{row}'] = description
        row += 1

    row += 2
    ws[f'A{row}'] = "Calculation Formulas"
    apply_style(ws[f'A{row}'], create_header_style())
    ws.merge_cells(f'A{row}:C{row}')

    formulas = [
        ("Regional Adjustment", "base_cost × regional_modifier", "See 04_Regional_Modifiers"),
        ("Timeline Premium", "base_cost × (1 + premium_rate)", "15% for accelerated schedules"),
        ("Total Cost", "Σ(categories) × regional_modifier × timeline_factor", "Compound adjustments")
    ]

    row += 1
    for formula_name, formula, note in formulas:
        ws[f'A{row}'] = formula_name
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = formula
        ws[f'B{row}'].font = Font(name='Courier New')
        ws[f'C{row}'] = note
        ws[f'C{row}'].font = Font(italic=True)
        row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 45
    ws.column_dimensions['C'].width = 35

def create_agent_instructions(ws, store_info):
    """Create agent instructions worksheet."""
    ws['A1'] = "AI Agent Usage Instructions"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')

    instructions = [
        ("Purpose", f"This budget plan provides a complete cost example for {store_info['name']} stores based on historical project data."),
        ("", ""),
        ("How to Use", "1. Review Executive Summary for high-level cost breakdown"),
        ("", "2. Reference Detailed Line Items for specific cost components"),
        ("", "3. Compare Scenario Comparisons to understand cost impacts of variations"),
        ("", "4. Cite Data Sources when referencing this budget in agent responses"),
        ("", ""),
        ("Adjusting for New Projects", "Parameters to customize:"),
        ("", "- Square footage (multiply line items proportionally)"),
        ("", "- Regional market (apply regional_modifiers from folder 04)"),
        ("", "- Timeline (apply 15% premium for accelerated schedules)"),
        ("", "- Constraints (reference folder 02 for cost impacts)"),
        ("", ""),
        ("Validation Checkpoints", "Before using this budget plan:"),
        ("", "✓ Confirm store type matches project requirements"),
        ("", "✓ Verify square footage is within typical range"),
        ("", "✓ Check regional modifiers are current"),
        ("", "✓ Review recent vendor negotiations for pricing updates"),
        ("", ""),
        ("Source Attribution", "When citing this budget in responses:"),
        ("", f"'Based on Budget_Plan_{store_info['type_id']}_{store_info['typical_sqft']}sqft.xlsx"),
        ("", f" which analyzes historical {store_info['name']} store costs'"),
        ("", ""),
        ("Named Ranges", "Key data zones (for programmatic access):"),
        ("", "- CostSummary: Executive Summary!B6:D13"),
        ("", "- LineItemsData: Detailed Line Items!A4:G[last_row]"),
        ("", "- ScenarioCompare: Scenario Comparisons!A4:E9"),
        ("", ""),
        ("Updates", "To maintain accuracy:"),
        ("", "- Regenerate quarterly as new projects complete"),
        ("", "- Update vendor pricing after negotiations"),
        ("", "- Adjust regional modifiers annually")
    ]

    row = 3
    for label, text in instructions:
        if label:
            ws[f'A{row}'] = label
            ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = text
        if text.startswith("✓"):
            ws[f'B{row}'].font = Font(color='00B050')
        row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 70

# Generate budget plans for each store type
for store_type in store_types_data['store_types']:
    create_budget_plan(store_type['type_id'])

print(f"  ✓ Created {len(store_types_data['store_types'])} budget plan files")

# ============================================================================
# 2. CONFIGURATION TEMPLATES
# ============================================================================
print("\n[2/4] Generating configuration templates...")

def create_store_configuration_template():
    """Create reusable store configuration template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Configuration Input"

    # Header
    ws['A1'] = "Store Build Configuration Template"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')

    ws['A2'] = "Agent Input Zone - Enter project parameters below"
    ws['A2'].font = Font(italic=True, color='0066CC')
    ws.merge_cells('A2:D2')

    # Input section
    row = 4
    ws[f'A{row}'] = "INPUT PARAMETERS"
    apply_style(ws[f'A{row}'], create_header_style())
    ws.merge_cells(f'A{row}:B{row}')

    inputs = [
        ("Store Type", "suburban_standard", "Options: urban_flagship, suburban_standard, express_compact, remodel_refresh, prototype_innovation"),
        ("Square Footage", 3500, "Enter store size in square feet"),
        ("Region/Market", "Columbus", "Enter city name (see Regional Modifiers list)"),
        ("Timeline", "Standard", "Options: Standard (12 weeks), Accelerated (8 weeks), Extended (16+ weeks)"),
        ("Special Constraints", "None", "Optional: landlord restrictions, budget caps, etc.")
    ]

    row += 1
    for label, default, note in inputs:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = default
        apply_style(ws[f'B{row}'], create_input_style())
        ws[f'C{row}'] = note
        ws[f'C{row}'].font = Font(italic=True, size=9)
        row += 1

    # Auto-calculation section
    row += 2
    ws[f'A{row}'] = "AUTO-CALCULATED BUDGET"
    apply_style(ws[f'A{row}'], create_header_style())
    ws.merge_cells(f'A{row}:D{row}')

    row += 1
    ws[f'A{row}'] = "Category"
    ws[f'B{row}'] = "Base Cost"
    ws[f'C{row}'] = "Regional Mult."
    ws[f'D{row}'] = "Adjusted Cost"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], create_header_style())

    # Placeholder formulas (would reference lookup tables)
    categories = ['Construction', 'Electrical', 'HVAC', 'Plumbing', 'Fixtures', 'Technology', 'Soft Costs']
    row += 1
    for cat in categories:
        ws[f'A{row}'] = cat
        ws[f'B{row}'] = "[FORMULA]"
        apply_style(ws[f'B{row}'], create_formula_style())
        ws[f'C{row}'] = "[LOOKUP]"
        apply_style(ws[f'C{row}'], create_formula_style())
        ws[f'D{row}'] = "[B * C]"
        apply_style(ws[f'D{row}'], create_output_style())
        row += 1

    # Total
    ws[f'A{row}'] = "TOTAL PROJECT COST"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws[f'D{row}'] = "[SUM]"
    ws[f'D{row}'].font = Font(bold=True, size=12)
    apply_style(ws[f'D{row}'], create_output_style())

    # Instructions worksheet
    ws_inst = wb.create_sheet("How to Use")
    ws_inst['A1'] = "Template Usage Instructions"
    ws_inst['A1'].font = Font(bold=True, size=14)

    instructions = [
        "1. Enter project parameters in the blue INPUT PARAMETERS cells",
        "2. The AUTO-CALCULATED BUDGET section will update automatically",
        "3. Review the calculated costs against historical averages",
        "4. Export results or reference in agent response",
        "",
        "For AI Agents:",
        "- Named range 'InputParameters' contains cells B5:B9",
        "- Named range 'BudgetOutput' contains cells D14:D21",
        "- Use VLOOKUP formulas to reference cost_model and regional_modifiers data",
        "",
        "Validation:",
        "- Store Type must match one of the 5 defined types",
        "- Square Footage should be within typical range for store type",
        "- Region must exist in regional_modifiers.csv"
    ]

    for idx, instruction in enumerate(instructions, start=3):
        ws_inst[f'A{idx}'] = instruction
        if instruction.startswith("For AI") or instruction.startswith("Validation"):
            ws_inst[f'A{idx}'].font = Font(bold=True)

    ws_inst.column_dimensions['A'].width = 70

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 18
    ws.column_dimensions['D'].width = 20

    filename = f"{output_dir}/templates/Template_Config_Store_Configuration.xlsx"
    wb.save(filename)
    print(f"  ✓ Created {filename}")

def create_constraint_response_template():
    """Create constraint response template."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Constraint Impact"

    ws['A1'] = "Constraint Response Template"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:D1')

    row = 4
    ws[f'A{row}'] = "SELECT CONSTRAINT TYPE"
    apply_style(ws[f'A{row}'], create_header_style())
    ws.merge_cells(f'A{row}:B{row}')

    row += 1
    ws[f'A{row}'] = "Constraint Type:"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = "[DROPDOWN]"
    apply_style(ws[f'B{row}'], create_input_style())
    ws[f'C{row}'] = "Options: landlord, budget, timeline, regional, operational"
    ws[f'C{row}'].font = Font(italic=True, size=9)

    row += 2
    ws[f'A{row}'] = "COST IMPACT CALCULATION"
    apply_style(ws[f'A{row}'], create_header_style())
    ws.merge_cells(f'A{row}:D{row}')

    row += 1
    ws[f'A{row}'] = "Impact Factor"
    ws[f'B{row}'] = "Affected Categories"
    ws[f'C{row}'] = "Multiplier"
    ws[f'D{row}'] = "Mitigation Strategy"
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws[f'{col}{row}'], create_header_style())

    # Example impacts
    impacts = [
        ("Approved vendor list", "All categories", "1.05 - 1.15", "Negotiate vendor approval"),
        ("No structural changes", "Construction", "1.00 (no change)", "Design within existing footprint"),
        ("After-hours work", "Labor categories", "1.20 - 1.30", "Optimize scheduling"),
        ("Accelerated timeline", "All categories", "1.15", "Parallel work streams")
    ]

    row += 1
    for impact, categories, mult, mitigation in impacts:
        ws[f'A{row}'] = impact
        ws[f'B{row}'] = categories
        ws[f'C{row}'] = mult
        ws[f'C{row}'].alignment = Alignment(horizontal='center')
        ws[f'D{row}'] = mitigation
        row += 1

    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 30

    filename = f"{output_dir}/templates/Template_Config_Constraint_Response.xlsx"
    wb.save(filename)
    print(f"  ✓ Created {filename}")

create_store_configuration_template()
create_constraint_response_template()

print(f"  ✓ Created 2 configuration template files")

# ============================================================================
# 3. BUILD STRATEGY WORKSHEETS
# ============================================================================
print("\n[3/4] Generating build strategy worksheets...")

def create_fast_track_strategy():
    """Create fast-track build strategy worksheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Fast Track Strategy"

    ws['A1'] = "Fast-Track Build Strategy"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')

    ws['A2'] = "Timeline Compression Techniques and Cost Impacts"
    ws.merge_cells('A2:E2')

    row = 4
    ws[f'A{row}'] = "STRATEGY OVERVIEW"
    apply_style(ws[f'A{row}'], create_header_style())
    ws.merge_cells(f'A{row}:E{row}')

    overview = [
        ("Objective", "Reduce 12-week standard timeline to 8 weeks"),
        ("Primary Driver", "Accelerated store opening for critical locations"),
        ("Cost Impact", "+15% to +20% total project cost"),
        ("Risk Level", "Medium-High (productivity loss, quality control)"),
        ("Success Rate", "78% on-time completion (vs. 87% standard)")
    ]

    row += 1
    for label, value in overview:
        ws[f'A{row}'] = label
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'B{row}'] = value
        ws.merge_cells(f'B{row}:E{row}')
        row += 1

    # Cost impact table
    row += 2
    ws[f'A{row}'] = "COST IMPACT BY CATEGORY"
    apply_style(ws[f'A{row}'], create_header_style())
    ws.merge_cells(f'A{row}:E{row}')

    row += 1
    ws[f'A{row}'] = "Category"
    ws[f'B{row}'] = "Standard Cost"
    ws[f'C{row}'] = "Premium"
    ws[f'D{row}'] = "Fast-Track Cost"
    ws[f'E{row}'] = "Justification"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], create_header_style())

    # Example costs (based on suburban_standard)
    base_costs = {
        'Construction': 226625,
        'Electrical': 77700,
        'HVAC': 51800,
        'Plumbing': 32375,
        'Fixtures': 161875,
        'Technology': 51800,
        'Soft Costs': 45325
    }

    premiums = {
        'Construction': 0.18,
        'Electrical': 0.15,
        'HVAC': 0.12,
        'Plumbing': 0.15,
        'Fixtures': 0.08,
        'Technology': 0.05,
        'Soft Costs': 0.10
    }

    justifications = {
        'Construction': "Overtime labor, weekend shifts",
        'Electrical': "Premium electrician rates, expedited materials",
        'HVAC': "Expedited delivery, installation coordination",
        'Plumbing': "Premium labor rates for compressed schedule",
        'Fixtures': "Air freight for fixtures, rush orders",
        'Technology': "Standard pricing (minimal timeline impact)",
        'Soft Costs': "Expedited permitting, additional oversight"
    }

    row += 1
    for cat, base_cost in base_costs.items():
        premium = premiums[cat]
        fast_track = int(base_cost * (1 + premium))

        ws[f'A{row}'] = cat
        ws[f'B{row}'] = base_cost
        ws[f'B{row}'].number_format = '$#,##0'
        ws[f'C{row}'] = premium
        ws[f'C{row}'].number_format = '0.0%'
        ws[f'D{row}'] = fast_track
        ws[f'D{row}'].number_format = '$#,##0'
        apply_style(ws[f'D{row}'], create_output_style())
        ws[f'E{row}'] = justifications[cat]
        ws[f'E{row}'].font = Font(size=9)
        row += 1

    # Total
    total_base = sum(base_costs.values())
    total_fast = sum(int(base_costs[cat] * (1 + premiums[cat])) for cat in base_costs)

    ws[f'A{row}'] = "TOTAL"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = total_base
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'C{row}'] = (total_fast - total_base) / total_base
    ws[f'C{row}'].number_format = '+0.0%'
    ws[f'C{row}'].font = Font(bold=True, color='C00000')
    ws[f'D{row}'] = total_fast
    ws[f'D{row}'].number_format = '$#,##0'
    ws[f'D{row}'].font = Font(bold=True)

    # Applicability criteria
    row += 3
    ws[f'A{row}'] = "WHEN TO USE THIS STRATEGY"
    apply_style(ws[f'A{row}'], create_header_style())
    ws.merge_cells(f'A{row}:E{row}')

    criteria = [
        "✓ Critical market entry timing (seasonal, competitive)",
        "✓ Lease penalty clauses for delayed opening",
        "✓ Budget can absorb 15-20% premium",
        "✓ Vendor capacity available for compressed schedule",
        "✓ Permitting expedite services available in jurisdiction",
        "✗ Complex site conditions (use standard timeline)",
        "✗ Landlord restrictions on after-hours work",
        "✗ Tight budget constraints (consider phased approach instead)"
    ]

    row += 1
    for criterion in criteria:
        ws[f'A{row}'] = criterion
        ws.merge_cells(f'A{row}:E{row}')
        if criterion.startswith("✓"):
            ws[f'A{row}'].font = Font(color='00B050')
        elif criterion.startswith("✗"):
            ws[f'A{row}'].font = Font(color='C00000')
        row += 1

    # Historical references
    row += 2
    ws[f'A{row}'] = "HISTORICAL PROJECT EXAMPLES"
    apply_style(ws[f'A{row}'], create_header_style())
    ws.merge_cells(f'A{row}:E{row}')

    row += 1
    ws[f'A{row}'] = "Store ID"
    ws[f'B{row}'] = "Market"
    ws[f'C{row}'] = "Actual Timeline"
    ws[f'D{row}'] = "Cost Premium"
    ws[f'E{row}'] = "Outcome"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], create_header_style())

    examples = [
        ("Store-147", "Columbus", "8.5 weeks", "17%", "Success - opened on time"),
        ("Store-162", "Cincinnati", "9 weeks", "19%", "Delayed 1 week due to permit"),
        ("Store-183", "Cleveland", "8 weeks", "16%", "Success - under budget"),
    ]

    row += 1
    for store, market, timeline, premium, outcome in examples:
        ws[f'A{row}'] = store
        ws[f'B{row}'] = market
        ws[f'C{row}'] = timeline
        ws[f'D{row}'] = premium
        ws[f'E{row}'] = outcome
        ws[f'E{row}'].font = Font(size=9)
        row += 1

    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 40

    filename = f"{output_dir}/strategy_worksheets/Strategy_Fast_Track.xlsx"
    wb.save(filename)
    print(f"  ✓ Created {filename}")

def create_value_engineering_strategy():
    """Create value engineering strategy worksheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Value Engineering"

    ws['A1'] = "Value Engineering Strategy"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')

    ws['A2'] = "Cost Reduction Opportunities by Category"
    ws.merge_cells('A2:E2')

    row = 4
    ws[f'A{row}'] = "COST REDUCTION OPPORTUNITIES"
    apply_style(ws[f'A{row}'], create_header_style())
    ws.merge_cells(f'A{row}:E{row}')

    row += 1
    ws[f'A{row}'] = "Category"
    ws[f'B{row}'] = "Standard Cost"
    ws[f'C{row}'] = "VE Approach"
    ws[f'D{row}'] = "Savings"
    ws[f'E{row}'] = "Quality Impact"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], create_header_style())

    opportunities = [
        ("Flooring", 63000, "LVT → Laminate", 0.22, "Minimal - 3yr vs 5yr lifespan"),
        ("Lighting Fixtures", 26780, "Reduce fixture count by 15%", 0.15, "Minor - adequate illumination maintained"),
        ("Mannequins", 6930, "Reduce count from 18 to 12", 0.33, "Minimal - display flexibility"),
        ("Digital Displays", 8600, "Reduce from 4 to 2 units", 0.50, "Moderate - less dynamic content"),
        ("Paint Finish", 15750, "Standard vs premium", 0.18, "Minimal - appearance"),
        ("Dressing Rooms", 14800, "Reduce from 8 to 6 units", 0.25, "Minimal - adequate capacity")
    ]

    total_standard = 0
    total_savings = 0

    row += 1
    for cat, standard, approach, savings_pct, impact in opportunities:
        savings = int(standard * savings_pct)
        ve_cost = standard - savings

        ws[f'A{row}'] = cat
        ws[f'B{row}'] = standard
        ws[f'B{row}'].number_format = '$#,##0'
        ws[f'C{row}'] = approach
        ws[f'C{row}'].font = Font(size=9)
        ws[f'D{row}'] = savings
        ws[f'D{row}'].number_format = '$#,##0'
        ws[f'D{row}'].font = Font(color='00B050', bold=True)
        ws[f'E{row}'] = impact
        ws[f'E{row}'].font = Font(size=9)

        total_standard += standard
        total_savings += savings
        row += 1

    # Total
    ws[f'A{row}'] = "TOTAL SAVINGS POTENTIAL"
    ws[f'A{row}'].font = Font(bold=True)
    ws[f'B{row}'] = total_standard
    ws[f'B{row}'].number_format = '$#,##0'
    ws[f'B{row}'].font = Font(bold=True)
    ws[f'D{row}'] = total_savings
    ws[f'D{row}'].number_format = '$#,##0'
    ws[f'D{row}'].font = Font(bold=True, color='00B050')
    ws[f'E{row}'] = f"{total_savings/total_standard:.1%} total reduction"
    ws[f'E{row}'].font = Font(bold=True)

    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 35

    filename = f"{output_dir}/strategy_worksheets/Strategy_Value_Engineering.xlsx"
    wb.save(filename)
    print(f"  ✓ Created {filename}")

create_fast_track_strategy()
create_value_engineering_strategy()

print(f"  ✓ Created 2 build strategy worksheet files")

# ============================================================================
# 4. AGENT TOOLS
# ============================================================================
print("\n[4/4] Generating agent integration tools...")

def create_master_index():
    """Create master index of all budget artifacts."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Master Catalog"

    ws['A1'] = "Budget Artifacts Master Index"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:E1')

    ws['A2'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws.merge_cells('A2:E2')

    # Budget Plans
    row = 4
    ws[f'A{row}'] = "BUDGET PLANS"
    apply_style(ws[f'A{row}'], create_header_style())
    ws.merge_cells(f'A{row}:E{row}')

    row += 1
    ws[f'A{row}'] = "File Name"
    ws[f'B{row}'] = "Store Type"
    ws[f'C{row}'] = "Size"
    ws[f'D{row}'] = "Use Cases"
    ws[f'E{row}'] = "Avg Cost"
    for col in ['A', 'B', 'C', 'D', 'E']:
        apply_style(ws[f'{col}{row}'], create_header_style())

    row += 1
    for store in store_types_data['store_types']:
        hist_avg = calculate_historical_average(store['type_id'])

        ws[f'A{row}'] = f"Budget_Plan_{store['type_id']}_{store['typical_sqft']}sqft.xlsx"
        ws[f'B{row}'] = store['name']
        ws[f'C{row}'] = f"{store['typical_sqft']:,} sqft"
        ws[f'D{row}'] = ", ".join(store['typical_use_cases'][:2])
        ws[f'D{row}'].font = Font(size=9)
        if hist_avg:
            ws[f'E{row}'] = hist_avg['total_cost']
            ws[f'E{row}'].number_format = '$#,##0'
        row += 1

    # Templates
    row += 2
    ws[f'A{row}'] = "CONFIGURATION TEMPLATES"
    apply_style(ws[f'A{row}'], create_header_style())
    ws.merge_cells(f'A{row}:E{row}')

    templates_list = [
        ("Template_Config_Store_Configuration.xlsx", "Universal store configuration input", "All store types"),
        ("Template_Config_Constraint_Response.xlsx", "Constraint impact calculator", "Projects with constraints")
    ]

    row += 1
    for filename, description, use_case in templates_list:
        ws[f'A{row}'] = filename
        ws[f'B{row}'] = description
        ws.merge_cells(f'B{row}:D{row}')
        ws[f'E{row}'] = use_case
        ws[f'E{row}'].font = Font(size=9)
        row += 1

    # Strategies
    row += 2
    ws[f'A{row}'] = "BUILD STRATEGIES"
    apply_style(ws[f'A{row}'], create_header_style())
    ws.merge_cells(f'A{row}:E{row}')

    strategies_list = [
        ("Strategy_Fast_Track.xlsx", "Timeline compression (8 weeks)", "+15-20% cost", "Critical timing"),
        ("Strategy_Value_Engineering.xlsx", "Cost reduction techniques", "-12-18% savings", "Budget constraints")
    ]

    row += 1
    for filename, description, impact, when in strategies_list:
        ws[f'A{row}'] = filename
        ws[f'B{row}'] = description
        ws[f'C{row}'] = impact
        ws[f'D{row}'] = when
        ws.merge_cells(f'D{row}:E{row}')
        row += 1

    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 18

    # Recommendation Matrix worksheet
    ws_rec = wb.create_sheet("Recommendation Matrix")
    ws_rec['A1'] = "Budget Artifact Recommendation Matrix"
    ws_rec['A1'].font = Font(bold=True, size=14)
    ws_rec.merge_cells('A1:D1')

    ws_rec['A3'] = "IF project has..."
    ws_rec['B3'] = "AND constraints include..."
    ws_rec['C3'] = "THEN use..."
    ws_rec['D3'] = "Plus consider..."
    for col in ['A', 'B', 'C', 'D']:
        apply_style(ws_rec[f'{col}3'], create_header_style())

    recommendations = [
        ("New urban flagship", "Premium finish requirement", "Budget_Plan_urban_flagship_5000sqft.xlsx", "None"),
        ("Standard suburban store", "No special constraints", "Budget_Plan_suburban_standard_3500sqft.xlsx", "None"),
        ("Any store type", "Tight timeline (< 10 weeks)", "Appropriate budget plan", "Strategy_Fast_Track.xlsx"),
        ("Any store type", "Budget cap constraint", "Appropriate budget plan", "Strategy_Value_Engineering.xlsx"),
        ("Unknown configuration", "Need to explore options", "Template_Config_Store_Configuration.xlsx", "Multiple budget plans"),
        ("Remodel project", "Store must stay open", "Budget_Plan_remodel_refresh_3500sqft.xlsx", "Constraint_Response template")
    ]

    row = 4
    for project, constraints, primary, secondary in recommendations:
        ws_rec[f'A{row}'] = project
        ws_rec[f'B{row}'] = constraints
        ws_rec[f'C{row}'] = primary
        ws_rec[f'C{row}'].font = Font(bold=True, color='0066CC')
        ws_rec[f'D{row}'] = secondary
        ws_rec[f'D{row}'].font = Font(italic=True)
        row += 1

    ws_rec.column_dimensions['A'].width = 25
    ws_rec.column_dimensions['B'].width = 30
    ws_rec.column_dimensions['C'].width = 40
    ws_rec.column_dimensions['D'].width = 35

    filename = f"{output_dir}/agent_tools/Master_Index_Budget_Artifacts.xlsx"
    wb.save(filename)
    print(f"  ✓ Created {filename}")

def create_sample_workflow():
    """Create sample agent workflow demonstration."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Sample Workflow"

    ws['A1'] = "Agent Workflow: Generate New Store Budget"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:C1')

    ws['A2'] = "Step-by-step example of using budget artifacts"
    ws.merge_cells('A2:C2')

    workflow_steps = [
        ("STEP 1: Receive Request", "", ""),
        ("User Request", "Create budget estimate for 3,200 sqft suburban store in Cincinnati with 10-week timeline", ""),
        ("", "", ""),
        ("STEP 2: Identify Parameters", "", ""),
        ("Store Type", "suburban_standard (closest match to 3,200 sqft)", "Extracted from request"),
        ("Region", "Cincinnati", "Extracted from request"),
        ("Timeline", "Accelerated (10 weeks vs standard 12)", "Derived constraint"),
        ("Square Footage", "3,200 sqft", "Exact specification"),
        ("", "", ""),
        ("STEP 3: Select Budget Artifacts", "", ""),
        ("Primary Reference", "Budget_Plan_suburban_standard_3500sqft.xlsx", "Base cost model"),
        ("Regional Adjustment", "regional_modifiers.csv (Cincinnati)", "1.06x multiplier"),
        ("Timeline Adjustment", "Strategy_Fast_Track.xlsx", "+15% premium"),
        ("", "", ""),
        ("STEP 4: Calculate Adjusted Budget", "", ""),
        ("Base Cost (3500 sqft)", "$647,500", "From budget plan"),
        ("Size Adjustment (3200/3500)", "$592,914", "Proportional scaling"),
        ("Cincinnati Multiplier (1.06)", "$628,489", "Regional adjustment"),
        ("Timeline Premium (1.15)", "$722,762", "Accelerated schedule"),
        ("FINAL ESTIMATED COST", "$722,762", "Agent output"),
        ("", "", ""),
        ("STEP 5: Cite Sources", "", ""),
        ("Agent Response", "Based on Budget_Plan_suburban_standard_3500sqft.xlsx", ""),
        ("", "adjusted for Cincinnati market (1.06x regional modifier)", ""),
        ("", "and 10-week accelerated timeline (+15% premium)", ""),
        ("", "per Strategy_Fast_Track.xlsx", ""),
        ("", "", ""),
        ("STEP 6: Provide Breakdown", "", ""),
        ("Construction", "$296,423", "35% of total"),
        ("Electrical", "$101,484", "14% of total"),
        ("HVAC", "$53,795", "7.5% of total"),
        ("Plumbing", "$33,621", "4.7% of total"),
        ("Fixtures", "$168,064", "23.3% of total"),
        ("Technology", "$53,795", "7.5% of total"),
        ("Soft Costs", "$47,076", "6.5% of total"),
        ("Contingency (10%)", "$65,705", "Included in soft costs")
    ]

    row = 4
    for col1, col2, col3 in workflow_steps:
        ws[f'A{row}'] = col1
        ws[f'B{row}'] = col2
        ws[f'C{row}'] = col3

        # Style step headers
        if col1.startswith("STEP"):
            ws[f'A{row}'].font = Font(bold=True, size=12, color='FFFFFF')
            ws[f'A{row}'].fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            ws.merge_cells(f'A{row}:C{row}')

        # Style key results
        if col1 in ["Base Cost (3500 sqft)", "Size Adjustment (3200/3500)", "Cincinnati Multiplier (1.06)", "Timeline Premium (1.15)"]:
            ws[f'B{row}'].font = Font(bold=True)

        if col1 == "FINAL ESTIMATED COST":
            ws[f'A{row}'].font = Font(bold=True, size=12)
            ws[f'B{row}'].font = Font(bold=True, size=12, color='FFFFFF')
            ws[f'B{row}'].fill = PatternFill(start_color='00B050', end_color='00B050', fill_type='solid')

        # Style category breakdown
        if col1 in ["Construction", "Electrical", "HVAC", "Plumbing", "Fixtures", "Technology", "Soft Costs", "Contingency (10%)"]:
            ws[f'A{row}'].font = Font(italic=True)
            ws[f'B{row}'].number_format = '$#,##0'
            ws[f'C{row}'].font = Font(italic=True, size=9)

        row += 1

    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 25

    filename = f"{output_dir}/agent_tools/Sample_Workflow_Demo.xlsx"
    wb.save(filename)
    print(f"  ✓ Created {filename}")

create_master_index()
create_sample_workflow()

print(f"  ✓ Created 2 agent tool files")

# ============================================================================
# SUMMARY
# ============================================================================
print("\n" + "="*60)
print("BUDGET ARTIFACTS GENERATION COMPLETE")
print("="*60)
print(f"\n08_Budget_Artifacts/")
print(f"  budget_plans/          {len(store_types_data['store_types'])} store-type budget plans")
print(f"  templates/             2 reusable configuration templates")
print(f"  strategy_worksheets/   2 build strategy worksheets")
print(f"  agent_tools/           2 agent integration tools")
print(f"\nTotal Excel files: {len(store_types_data['store_types']) + 2 + 2 + 2}")
print("="*60 + "\n")
